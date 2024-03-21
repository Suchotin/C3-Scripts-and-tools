import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import getpass
import re

def find_excels(root_folder, date_range, days):
    file_paths = []

    def check_modification_time(modification_time):
        if date_range:
            return start_date <= modification_time <= end_date
        elif days:
            return datetime.now() - modification_time <= timedelta(days=days)
        return False

    if date_range is not None:
        start_date, end_date = date_range
    elif days is not None:
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)

    for root, dirs, files in os.walk(root_folder):
        # Check if the current directory is a second-level subdirectory
        if os.path.relpath(root, root_folder).count(os.sep) == 1:
            for file in files:
                file_path = os.path.join(root, file)
                modification_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                if file.lower().endswith('.xlsx') and file.lower().startswith('offer') and check_modification_time(modification_time):
                    file_paths.append(file_path)

    return file_paths
def read_excel_data(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]

    art = []
    name = []
    count = []
    units = []
    price_final = []
    CO_date = []

    global digits
    digits = None

    max_col = ws.max_column
    max_row = ws.max_row

    com_col = None
    res_row = None

    # Find the column with 'Примечания'
    for col in range(3, max_col + 1):
        if ws.cell(15, col).value == 'Примечания':
            com_col = col
            pf_col = com_col - 1
            break

    # Find the row with 'ИТОГО:'
    if com_col:
        for col in range(3, com_col):
            for row in range(15, max_row + 1):
                if ws.cell(row, col).value == 'ИТОГО:':
                    res_row = row
                    break
            if res_row:
                break

    e8value = ws['E8'].value

    # Extract digits using regular expression
    if e8value:
        match = re.search(r'№ (\d+)\s*', str(e8value))
        digits = int(match.group(1)) if match else None

    # Extract date using regular expression
    extracted_date = re.search(r'\b(\d{2}\.\d{2}\.\d{4})\b', e8value).group(1) if e8value else None

    # Read data from Excel
    if com_col and res_row:
        for row in range(16, res_row):
            art.append(ws.cell(row, 3).value)
            name.append(ws.cell(row, 4).value)
            count.append(ws.cell(row, 5).value)
            units.append(ws.cell(row, 6).value)
            price_final.append(ws.cell(row, pf_col).value)
            CO_date.append(extracted_date)

    count = [int(value) if value is not None and type(value) is not str else 0 for value in count]
    data = np.array([art, name, count, units, price_final, CO_date]).T.tolist()
    data_f = list(filter(lambda x: x[4] is not None, data))

    # Convert data to DataFrame
    columns = ['Артикул', 'Наименование', 'Количество', 'Единица измерения', 'Финальная стоимость, руб.', 'Дата модификации']
    index = [digits] * len(data_f) if digits is not None else [None] * len(data_f)
    excel_data = pd.DataFrame(data_f, columns=columns, index=index)

    # Set display options
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    pd.set_option('display.max_colwidth', None)

    # Insert Presale column
    username = getpass.getuser()
    excel_data.insert(0, 'Presale', username)

    return digits, excel_data

def update_dataset(csv_file, new_data):
    try:
        if os.path.exists(csv_file):
            # Read existing dataset from CSV
            df = pd.read_csv(csv_file, index_col='Project_ID')
        else:
            # Create a new DataFrame if CSV file doesn't exist
            df = pd.DataFrame()

        # Check if project ID exists in the dataset and delete corresponding rows
        if digits in df.index:
            df = df.drop(digits)

        # Append new data to the dataset
        df = pd.concat([df, new_data])
        df.index = df.index.rename('Project_ID')

        # Convert the index to integers
        df.index = df.index.astype(int)

        # Save updated dataset to CSV
        df.to_csv(csv_file)

        print(f"Dataset updated with project ID: {digits}")
    except Exception as e:
        print(f"Error updating dataset: {e}")

def parse_date_input(input_str):
    parts = input_str.split('-')
    if len(parts) == 2:
        try:
            start_date = datetime.strptime(parts[0].strip(), "%d.%m.%Y")
            end_date = datetime.strptime(parts[1].strip(), "%d.%m.%Y")
            return None, (start_date, end_date)
        except ValueError:
            pass
    else:
        try:
            days = int(input_str)
            return days, None
        except ValueError:
            pass
    return None, None

def main():
    # Define the root folder containing Excel files
    root_folder = rf'C:\Users\{getpass.getuser()}\YandexDisk\C3 Presale\=КП'

    # Ask the user for the time period in days or a date range
    while True:
        input_str = input("Enter the time period (days or date range 'dd.mm.yyyy - dd.mm.yyyy'): ")
        days, date_range = parse_date_input(input_str)
        if days is not None or date_range is not None:
            break
        print("Invalid input. Please enter a single integer (days) or a date range in the format 'dd.mm.yyyy - dd.mm.yyyy'.")

    excels_found = find_excels(root_folder, date_range=date_range, days=days)

    # Define CSV file for the dataset
    folder_path = rf"C:\Users\{getpass.getuser()}\YandexDisk\C3 Presale\Прочее\CO_DATABASE\CO_PARCER"
    csv_filename = f"dataset_{getpass.getuser()}.csv"
    dataset_csv = os.path.join(folder_path, csv_filename)

    # Read data from Excel files using openpyxl
    for excel_file in excels_found:

        # Parce one Excel file
        project_id, excel_data = read_excel_data(excel_file)

        # Update dataset based on project ID
        update_dataset(dataset_csv, excel_data)

        print(excel_data)


if __name__ == "__main__":
    main()
